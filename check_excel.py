import pandas as pd
import openpyxl

file_path = r'C:\dev\AppDSI\BONDECOMMANDE.xlsx'

wb = openpyxl.load_workbook(file_path)
sheet_names = wb.sheetnames
print('Onglets:', sheet_names)
print('='*80)

test_sheet = 'LASER MONO'
if test_sheet in sheet_names:
    ws = wb[test_sheet]
    print(f'\nOnglet: {test_sheet}')
    print('Premières 5 lignes:')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        print(f'Row {i}: {row}')

print('\n' + '='*80)
df = pd.read_excel(file_path, sheet_name='LASER MONO', header=None)
print('\nAvec Pandas (premières 5 lignes):')
print(df.head(5))
print(f'\nNombre de colonnes: {len(df.columns)}')
print(f'Nombre de lignes: {len(df)}')
