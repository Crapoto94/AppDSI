import openpyxl

file_path = r'C:\dev\AppDSI\BONDECOMMANDE.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['DIVERS']

print('Onglet DIVERS - Premières 8 lignes:')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
    print(f'Row {i}: {row}')

print('\nAvec XLSX utils:')
from openpyxl import load_workbook
import openpyxl.utils

wb = load_workbook(file_path)
ws = wb['DIVERS']

# Get all values
all_values = []
for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
    all_values.append(row)

for i, row in enumerate(all_values, 1):
    print(f'Index {i-1}: {row}')
